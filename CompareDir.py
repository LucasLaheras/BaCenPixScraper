import os
import pandas as pd
import warnings
import shutil

# Suprimir avisos do UserWarning
warnings.simplefilter("ignore", UserWarning)


def compare_files(new_dir, old_dir, output_dir):
    # Criar dicionários para armazenar arquivos com caminhos relativos
    files_dict_new = {}
    files_dict_old = {}

    # Percorrer o novo diretório e seus subdiretórios
    for root, dirs, files in os.walk(new_dir):
        for file in files:
            rel_dir = os.path.relpath(root, new_dir)
            rel_file = os.path.join(rel_dir, file)
            files_dict_new[rel_file] = os.path.join(root, file)

    # Percorrer o diretório antigo e seus subdiretórios
    for root, dirs, files in os.walk(old_dir):
        for file in files:
            rel_dir = os.path.relpath(root, old_dir)
            rel_file = os.path.join(rel_dir, file)
            files_dict_old[rel_file] = os.path.join(root, file)

    # Comparar arquivos com o mesmo caminho relativo entre os dois diretórios
    for rel_file in files_dict_new:
        output_file_path = os.path.join(output_dir, rel_file)

        new_file_path = files_dict_new[rel_file]

        if rel_file in files_dict_old:
            os.makedirs(os.path.dirname(output_file_path), exist_ok=True)

            old_file_path = files_dict_old[rel_file]
            if rel_file.endswith('.xlsx'):
                changed = compare_xlsx_files(new_file_path, old_file_path, output_file_path)

                if changed:
                    print(f"Arquivo {rel_file} foi alterado!")
            else:
                # Implementar outras comparações de arquivos, se necessário
                pass
        else:
            os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
            shutil.copy(new_file_path, output_file_path)
            print(f"Arquivo {rel_file} é novo no novo diretório.")

    os.makedirs(os.path.join(os.path.dirname(output_dir), "deleted"), exist_ok=True)

    # Arquivos que estão apenas no diretório antigo
    for rel_file in files_dict_old:
        output_file_path = os.path.join(output_dir, "deleted", rel_file)

        old_file_path = files_dict_old[rel_file]

        if rel_file not in files_dict_new:
            os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
            shutil.copy(old_file_path, output_file_path)

            print(f"Arquivo {rel_file} foi deletado no novo diretório.")


def compare_xlsx_files(new_file, old_file, output_file):
    changed = False

    # Carregar os arquivos novos e antigos em dataframes pandas
    new_df = pd.read_excel(new_file, sheet_name=None)
    old_df = pd.read_excel(old_file, sheet_name=None)

    # Criar um dicionário para armazenar as alterações
    changes = {}

    # Flag para indicar se há alterações
    has_changes = False

    # Iterar por cada planilha no novo arquivo
    for sheet_name, new_sheet in new_df.items():
        # Obter a planilha correspondente do arquivo antigo
        old_sheet = old_df.get(sheet_name)

        # Se a planilha existir no arquivo antigo, comparar os dados
        if old_sheet is not None:
            # Garantir que ambas as planilhas tenham as mesmas colunas
            common_cols = new_sheet.columns.intersection(old_sheet.columns)
            new_sheet = new_sheet[common_cols].copy()
            old_sheet = old_sheet[common_cols].copy()

            # Adicionar um identificador único se não estiver presente
            if 'ID' not in common_cols:
                new_sheet.reset_index(inplace=True)
                old_sheet.reset_index(inplace=True)
                id_col = 'index'
            else:
                id_col = 'ID'

            # Definir o identificador como o índice
            new_sheet.set_index(id_col, inplace=True)
            old_sheet.set_index(id_col, inplace=True)

            # Encontrar novas linhas
            new_rows = new_sheet.loc[~new_sheet.index.isin(old_sheet.index)].reset_index()
            if not new_rows.empty:
                has_changes = True
                new_rows['Change_Type'] = 'New'

            # Encontrar linhas atualizadas
            updated_rows = new_sheet.loc[new_sheet.index.isin(old_sheet.index)]
            old_rows = old_sheet.loc[old_sheet.index.isin(new_sheet.index)]

            # Handle NaNs and ensure consistent data types
            updated_rows_filled = updated_rows.fillna('<<NAN>>').astype(str)
            old_rows_filled = old_rows.fillna('<<NAN>>').astype(str)

            # Compute the changes mask
            changes_mask = ~(updated_rows_filled == old_rows_filled).all(axis=1)

            # Get the rows that have changes
            updated_rows = updated_rows.loc[changes_mask]
            old_rows = old_rows.loc[changes_mask]

            if not updated_rows.empty:
                has_changes = True
                # Combinar valores atualizados e antigos em um único DataFrame
                diff_df = updated_rows.copy()
                for col in updated_rows.columns:
                    diff_df[col + '_old'] = old_rows[col]
                    diff_df[col + '_new'] = updated_rows[col]
                    diff_df.drop(columns=col, inplace=True)
                diff_df.reset_index(inplace=True)
                diff_df['Change_Type'] = 'Updated'
            else:
                diff_df = pd.DataFrame()

            # Armazenar as alterações
            changes[sheet_name] = {
                'new_rows': new_rows,
                'updated_rows': diff_df
            }
        else:
            # Planilha inteira é nova
            new_sheet.reset_index(inplace=True)
            new_sheet['Change_Type'] = 'New'
            changes[sheet_name] = {
                'new_rows': new_sheet,
                'updated_rows': pd.DataFrame()
            }
            if not new_sheet.empty:
                has_changes = True

    # Planilhas que estão apenas no arquivo antigo (planilhas deletadas)
    deleted_sheets = set(old_df.keys()) - set(new_df.keys())
    for sheet_name in deleted_sheets:
        changes[sheet_name] = {
            'deleted_sheet': True,
            'content': old_df[sheet_name]
        }
        has_changes = True

    # Se não houver alterações, não criar o arquivo de saída
    if not has_changes:
        # print(f"Nenhuma alteração detectada entre {new_file} e {old_file}.")
        return False

    # Escrever as alterações no arquivo de saída
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, change in changes.items():
            if change.get('deleted_sheet'):
                # Indicar que a planilha foi deletada
                df = pd.DataFrame({'Mensagem': [f'A planilha "{sheet_name}" foi deletada.']})
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Combinar novas linhas e linhas atualizadas
                output_df = pd.concat([change['new_rows'], change['updated_rows']], ignore_index=True)
                # Modificação aqui: escrever a planilha mesmo se output_df estiver vazio
                output_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar formatação para destacar as diferenças
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(output_file)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for sheet_name in changes.keys():
        if changes[sheet_name].get('deleted_sheet'):
            continue
        # Adicionar verificação se a planilha existe no workbook
        if sheet_name not in wb.sheetnames:
            continue  # Pula para a próxima planilha
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Obter os índices das colunas de valores antigos e novos
        old_new_cols = []
        for i, header in enumerate(headers):
            if header.endswith('_old'):
                col_name = header[:-4]  # remover '_old'
                old_col_idx = i + 1  # Índices das colunas no Excel começam em 1
                new_col_name = col_name + '_new'
                if new_col_name in headers:
                    new_col_idx = headers.index(new_col_name) + 1
                    old_new_cols.append((old_col_idx, new_col_idx))

        # Índice da coluna 'Change_Type'
        if 'Change_Type' in headers:
            change_type_col_idx = headers.index('Change_Type') + 1
        else:
            continue

        # Iterar pelas linhas e comparar os valores antigos e novos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Verificar se a linha é 'Updated'
            change_type_cell = row[change_type_col_idx - 1]
            if change_type_cell.value == 'Updated':
                for old_col_idx, new_col_idx in old_new_cols:
                    old_cell = row[old_col_idx - 1]  # índice baseado em zero
                    new_cell = row[new_col_idx - 1]
                    if old_cell.value != new_cell.value:
                        # Aplicar preenchimento amarelo à célula nova
                        new_cell.fill = yellow_fill
                        changed = True

    wb.save(output_file)
    return changed

# Exemplo de uso da função compare_files
# output_dir = "/Users/lucaslaheras/PycharmProjects/BaCenDocumentation/test/"
#
# compare_files("/Users/lucaslaheras/PycharmProjects/BaCenDocumentation/temp/v5.09.1",
#               "/Users/lucaslaheras/PycharmProjects/BaCenDocumentation/temp/v5.08.1", output_dir)